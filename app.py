import io
import csv
import secrets
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, send_file)

from database import (
    init_db, assign_balanced_group, create_participant, get_participant,
    mark_participant_used, mark_participant_completed, save_response,
    get_all_responses, get_stats, generate_codes_batch, import_participants_csv,
    get_db, get_random_case, get_case, get_all_cases, create_case,
    update_case, delete_case, import_cases_csv, clear_all_data,
    get_setting, set_setting, reset_cases_to_default, get_collection_status,
    assign_cases_to_participant, get_current_participant_case,
    save_page1_answer, mark_participant_case_completed, count_participant_cases,
)

app = Flask(__name__)
app.secret_key = 'ctap-survey-secret-2026-change-in-prod'

ADMIN_PASSWORD = 'aiprojectgroup12'

with app.app_context():
    init_db()


# ──────────────────────────────────────────────
# Survey routes
# ──────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/start', methods=['POST'])
def start_survey():
    if not get_collection_status()['is_open']:
        return render_template('closed.html')
    group = assign_balanced_group()
    code = secrets.token_urlsafe(8)
    conn = get_db()
    while conn.execute('SELECT id FROM participants WHERE code = ?', (code,)).fetchone():
        code = secrets.token_urlsafe(8)
    conn.close()
    create_participant(code, group)
    return redirect(url_for('survey', code=code))


@app.route('/survey/<code>')
def survey(code):
    participant = get_participant(code)
    if not participant:
        return render_template('error.html', message='无效的访问码，请联系研究人员。'), 404
    if participant['completed']:
        return render_template('complete.html', already_done=True)

    # Block un-started participants when collection is closed
    if not participant['used'] and not get_collection_status()['is_open']:
        return render_template('closed.html')

    mark_participant_used(code)

    # Assign cases on first access
    total, done = count_participant_cases(code)
    if total == 0:
        n = int(get_setting('cases_per_participant', '1'))
        assign_cases_to_participant(code, n)
        total = n

    current = get_current_participant_case(code)
    if current is None:
        mark_participant_completed(code)
        return render_template('complete.html', already_done=False)

    group = current.get('case_group') or participant['group_name']
    return render_template(
        'survey.html',
        code=code,
        group=group,
        has_ai=(group in ('B', 'D')),
        has_timer=(group in ('C', 'D')),
        case=current,
        case_num=current['sequence_num'],
        total_cases=total,
    )


@app.route('/survey/<code>/page1', methods=['POST'])
def submit_page1(code):
    participant = get_participant(code)
    if not participant:
        return jsonify({'error': '无效的访问码'}), 404
    if participant['completed']:
        return jsonify({'error': '该问卷已提交'}), 400

    data = request.get_json(silent=True) or {}
    current = get_current_participant_case(code)
    if not current:
        return jsonify({'error': '没有待回答的情景'}), 400

    page1_answer = data.get('page1_answer')  # int 1/0 or None if timed out
    timed_out = bool(data.get('timed_out', False))
    try:
        page1_time = float(data['page1_time_seconds']) if data.get('page1_time_seconds') is not None else None
    except (TypeError, ValueError):
        page1_time = None

    save_page1_answer(
        code=code,
        sequence_num=current['sequence_num'],
        page1_answer=page1_answer,
        page1_time_seconds=page1_time,
        page1_timed_out=timed_out,
    )
    return jsonify({'success': True})


@app.route('/survey/<code>/submit', methods=['POST'])
def submit_survey(code):
    participant = get_participant(code)
    if not participant:
        return jsonify({'error': '无效的访问码'}), 404
    if participant['completed']:
        return jsonify({'error': '该问卷已提交', 'redirect': url_for('complete')}), 400

    data = request.get_json(silent=True) or {}

    current = get_current_participant_case(code)
    if not current:
        return jsonify({'error': '没有待提交的情景'}), 400

    group = current.get('case_group') or participant['group_name']
    has_ai = group in ('B', 'D')

    for field in ['responsibility', 'confidence', 'wait_for_intel']:
        if field not in data:
            return jsonify({'error': f'缺少必填字段: {field}'}), 400
    if has_ai and 'ai_influence' not in data:
        return jsonify({'error': '缺少必填字段: ai_influence'}), 400

    try:
        time_taken = float(data['time_taken_seconds'])
    except (KeyError, TypeError, ValueError):
        time_taken = None

    ai_prob_cond_str = current.get('ai_prob_condition')
    if has_ai and ai_prob_cond_str in ('high', 'low'):
        ai_prob_cond_int = 1 if ai_prob_cond_str == 'high' else 0
        prob_high = current.get('prob_high') or 0.85
        prob_low  = current.get('prob_low')  or 0.15
        ai_prob_val = prob_high if ai_prob_cond_str == 'high' else prob_low
    else:
        ai_prob_cond_int = None
        ai_prob_val = None

    save_response(
        code=code,
        group=group,
        approve_strike=current.get('page1_answer'),
        responsibility=int(data['responsibility']),
        confidence=int(data['confidence']),
        ai_influence=int(data['ai_influence']) if has_ai else None,
        wait_for_intel=int(data['wait_for_intel']),
        time_taken=time_taken,
        case_id=current['case_id'],
        case_title=current['title'],
        sequence_num=current['sequence_num'],
        page1_timed_out=current.get('page1_timed_out', 0),
        page1_time_seconds=current.get('page1_time_seconds'),
        timeout_reason=data.get('timeout_reason', '').strip() or None,
        ai_prob_condition=ai_prob_cond_int,
        ai_prob_value=ai_prob_val,
    )

    mark_participant_case_completed(code, current['sequence_num'])

    next_case = get_current_participant_case(code)
    if next_case:
        return jsonify({'success': True, 'next_case': True, 'redirect': url_for('survey', code=code)})

    mark_participant_completed(code)
    return jsonify({'success': True, 'next_case': False, 'redirect': url_for('complete')})


@app.route('/complete')
def complete():
    return render_template('complete.html', already_done=False)


# ──────────────────────────────────────────────
# Admin routes
# ──────────────────────────────────────────────

def admin_required():
    return session.get('admin_auth') is True


@app.route('/admin')
def admin():
    if admin_required():
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html', error=None)


@app.route('/admin/auth', methods=['POST'])
def admin_auth():
    if request.form.get('password', '') == ADMIN_PASSWORD:
        session['admin_auth'] = True
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html', error='密码错误，请重试。')


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_auth', None)
    return redirect(url_for('admin'))


@app.route('/admin/dashboard')
def admin_dashboard():
    if not admin_required():
        return redirect(url_for('admin'))
    stats = get_stats()
    cases_per_participant = int(get_setting('cases_per_participant', '1'))
    collection = get_collection_status()
    return render_template('admin_dashboard.html', stats=stats,
                           cases_per_participant=cases_per_participant,
                           collection=collection)


@app.route('/admin/settings', methods=['GET', 'POST'])
def admin_settings():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        cpp = data.get('cases_per_participant')
        if cpp is not None:
            try:
                cpp = max(1, min(int(cpp), 20))
                set_setting('cases_per_participant', cpp)
            except (TypeError, ValueError):
                return jsonify({'error': '无效数值'}), 400
        return jsonify({'success': True})
    return jsonify({'cases_per_participant': int(get_setting('cases_per_participant', '1'))})


@app.route('/admin/api/stats')
def admin_api_stats():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify(get_stats())


@app.route('/admin/api/responses')
def admin_api_responses():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify(get_all_responses())


@app.route('/admin/export/csv')
def export_csv():
    if not admin_required():
        return redirect(url_for('admin'))

    responses = get_all_responses()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ID', '参与者代码', '实验组', '情景序号',
        '是否批准打击(1=是/0=否/空=超时未答)', '第一页用时(秒)', '是否超时',
        '责任感(1-7)', '信心(1-7)', 'AI影响(1-7)', '是否等待更多情报(1=是/0=否)',
        'AI概率条件(1=高/0=低)', 'AI概率值',
        '超时原因', '总用时(秒)', '提交时间',
    ])
    for r in responses:
        writer.writerow([
            r['id'], r['participant_code'], r['group_name'], r.get('sequence_num', 1),
            r['approve_strike'] if r['approve_strike'] is not None else '超时未答',
            f"{r['page1_time_seconds']:.1f}" if r.get('page1_time_seconds') is not None else '',
            '是' if r.get('page1_timed_out') else '否',
            r['responsibility'], r['confidence'],
            r['ai_influence'] if r['ai_influence'] is not None else '',
            r['wait_for_intel'],
            r['ai_prob_condition'] if r.get('ai_prob_condition') is not None else '',
            f"{r['ai_prob_value']:.2f}" if r.get('ai_prob_value') is not None else '',
            r.get('timeout_reason') or '',
            f"{r['time_taken_seconds']:.1f}" if r['time_taken_seconds'] is not None else '',
            r['submitted_at'],
        ])

    output.seek(0)
    filename = f'survey_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/admin/export/excel')
def export_excel():
    if not admin_required():
        return redirect(url_for('admin'))

    responses = get_all_responses()
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = '问卷原始数据'
    hdr_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    headers = [
        'ID', '参与者代码', '实验组', '情景序号',
        '是否批准打击(1=是)', '第一页用时(秒)', '是否超时',
        '责任感(1-7)', '信心(1-7)', 'AI影响(1-7)', '是否等待更多情报(1=是)',
        'AI概率条件(1=高/0=低)', 'AI概率值',
        '超时原因', '总用时(秒)', '提交时间',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for ri, r in enumerate(responses, 2):
        ws.cell(ri, 1, r['id'])
        ws.cell(ri, 2, r['participant_code'])
        ws.cell(ri, 3, r['group_name'])
        ws.cell(ri, 4, r.get('sequence_num', 1))
        ws.cell(ri, 5, r['approve_strike'] if r['approve_strike'] is not None else '超时未答')
        ws.cell(ri, 6, round(r['page1_time_seconds'], 1) if r.get('page1_time_seconds') is not None else '')
        ws.cell(ri, 7, '是' if r.get('page1_timed_out') else '否')
        ws.cell(ri, 8, r['responsibility'])
        ws.cell(ri, 9, r['confidence'])
        ws.cell(ri, 10, r['ai_influence'] if r['ai_influence'] is not None else '')
        ws.cell(ri, 11, r['wait_for_intel'])
        ws.cell(ri, 12, r['ai_prob_condition'] if r.get('ai_prob_condition') is not None else '')
        ws.cell(ri, 13, round(r['ai_prob_value'], 2) if r.get('ai_prob_value') is not None else '')
        ws.cell(ri, 14, r.get('timeout_reason') or '')
        ws.cell(ri, 15, round(r['time_taken_seconds'], 1) if r['time_taken_seconds'] is not None else '')
        ws.cell(ri, 16, r['submitted_at'])

    ws2 = wb.create_sheet('分组统计')
    stats = get_stats()
    labels = {'A': '无AI + 低压力', 'B': '有AI + 低压力',
               'C': '无AI + 高压力', 'D': '有AI + 高压力'}
    sum_headers = ['实验组', '条件说明', '总人数', '批准打击人数', '批准率',
                   '平均责任感', '平均信心', '平均AI影响']
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for ri, group in enumerate(['A', 'B', 'C', 'D'], 2):
        d = stats['by_group'][group]
        ws2.cell(ri, 1, group); ws2.cell(ri, 2, labels[group])
        ws2.cell(ri, 3, d['count']); ws2.cell(ri, 4, d['approve_count'])
        ws2.cell(ri, 5, f"{d['approve_rate']:.1%}" if d['count'] > 0 else 'N/A')
        ws2.cell(ri, 6, d['avg_responsibility'] or 'N/A')
        ws2.cell(ri, 7, d['avg_confidence'] or 'N/A')
        ws2.cell(ri, 8, d['avg_ai_influence'] or 'N/A')

    for sheet in [ws, ws2]:
        for col in sheet.columns:
            width = max((len(str(cell.value or '')) for cell in col), default=8)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'survey_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/admin/generate-codes', methods=['POST'])
def admin_generate_codes():
    if not admin_required():
        return redirect(url_for('admin'))
    try:
        count = max(4, min(int(request.form.get('count', 20)), 1000))
    except ValueError:
        count = 20
    codes = generate_codes_batch(count)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['参与者代码', '实验组', '问卷链接'])
    for code, group in codes:
        survey_url = f"{request.host_url}survey/{code}"
        writer.writerow([code, group, survey_url])
    output.seek(0)
    filename = f'participant_codes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/admin/import', methods=['POST'])
def admin_import():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    if 'file' not in request.files:
        return jsonify({'error': '未选择文件'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': '未选择文件'}), 400
    if not file.filename.lower().endswith('.csv'):
        return jsonify({'error': '仅支持 CSV 文件'}), 400
    content = file.read().decode('utf-8-sig')
    result = import_participants_csv(content)
    return jsonify(result)


# ── Case management ────────────────────────────────

@app.route('/admin/cases')
def admin_cases():
    if not admin_required():
        return redirect(url_for('admin'))
    cases = get_all_cases()
    return render_template('admin_cases.html', cases=cases)


@app.route('/admin/api/cases')
def admin_api_cases():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify(get_all_cases())


@app.route('/admin/cases/add', methods=['POST'])
def admin_case_add():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    title = data.get('title', '').strip()
    scenario_text = data.get('scenario_text', '').strip()
    if not scenario_text:
        return jsonify({'error': '情景内容不能为空'}), 400
    if not title:
        title = f'情景 {datetime.now().strftime("%m%d%H%M")}'
    try:
        prob_high = max(0.0, min(float(data.get('prob_high', 85)), 100.0)) / 100.0
        prob_low  = max(0.0, min(float(data.get('prob_low',  15)), 100.0)) / 100.0
    except (TypeError, ValueError):
        prob_high, prob_low = 0.85, 0.15
    new_id = create_case(title, scenario_text, prob_high, prob_low)
    return jsonify({'success': True, 'id': new_id})


@app.route('/admin/cases/<int:case_id>', methods=['PUT'])
def admin_case_update(case_id):
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    title = data.get('title', '').strip()
    scenario_text = data.get('scenario_text', '').strip()
    active = data.get('active', True)
    if not scenario_text:
        return jsonify({'error': '情景内容不能为空'}), 400
    try:
        prob_high = max(0.0, min(float(data.get('prob_high', 85)), 100.0)) / 100.0
        prob_low  = max(0.0, min(float(data.get('prob_low',  15)), 100.0)) / 100.0
    except (TypeError, ValueError):
        prob_high, prob_low = 0.85, 0.15
    update_case(case_id, title, scenario_text, active, prob_high, prob_low)
    return jsonify({'success': True})


@app.route('/admin/cases/<int:case_id>', methods=['DELETE'])
def admin_case_delete(case_id):
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    delete_case(case_id)
    return jsonify({'success': True})


@app.route('/admin/cases/reset-default', methods=['POST'])
def admin_cases_reset_default():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    reset_cases_to_default()
    return jsonify({'success': True})


@app.route('/admin/cases/import', methods=['POST'])
def admin_cases_import():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    if 'file' not in request.files:
        return jsonify({'error': '未选择文件'}), 400
    file = request.files['file']
    if not file.filename or not file.filename.lower().endswith('.csv'):
        return jsonify({'error': '仅支持 CSV 文件'}), 400
    content = file.read().decode('utf-8-sig')
    result = import_cases_csv(content)
    return jsonify(result)


@app.route('/admin/collection', methods=['GET', 'POST'])
def admin_collection():
    if not admin_required():
        return redirect(url_for('admin')) if request.method == 'GET' else (jsonify({'error': 'Unauthorized'}), 401)
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        set_setting('collection_open',    '1' if data.get('manual_open', True) else '0')
        set_setting('collection_quota',    str(int(data['quota'])) if data.get('quota') else '')
        set_setting('collection_deadline', data.get('deadline', '').strip())
        return jsonify(get_collection_status())
    return render_template('admin_collection.html', collection=get_collection_status())


@app.route('/admin/clear-data', methods=['POST'])
def admin_clear_data():
    if not admin_required():
        return jsonify({'error': 'Unauthorized'}), 401
    if request.get_json(silent=True, force=True).get('password', '') != ADMIN_PASSWORD:
        return jsonify({'error': '密码错误'}), 403
    clear_all_data()
    return jsonify({'success': True})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
